from openpyxl import Workbook


def generate_excel(summary_data, room_output, inv_map):

    wb = Workbook()
    ws = wb.active
    ws.title = "SmartExam Seating"

    row_pointer = 1

    # ---------------- SUMMARY ---------------- #

    ws.cell(row=row_pointer, column=1, value="Total Rooms")
    ws.cell(row=row_pointer, column=2, value=summary_data["rooms"])
    row_pointer += 1

    ws.cell(row=row_pointer, column=1, value="Total Students")
    ws.cell(row=row_pointer, column=2, value=summary_data["students"])
    row_pointer += 1

    ws.cell(row=row_pointer, column=1, value="Total Capacity")
    ws.cell(row=row_pointer, column=2, value=summary_data["capacity"])
    row_pointer += 2

    # ---------------- INVIGILATORS ---------------- #

    ws.cell(row=row_pointer, column=1, value="Invigilator Allocation")
    row_pointer += 1

    for room, inv in inv_map.items():
        ws.cell(row=row_pointer, column=1, value=room)
        ws.cell(row=row_pointer, column=2, value=inv)
        row_pointer += 1

    row_pointer += 2

    # ---------------- SEATING ---------------- #

    for room_name, data in room_output.items():

        ws.cell(row=row_pointer, column=1, value=f"{room_name} Seating")
        row_pointer += 1

        ws.cell(row=row_pointer, column=1, value="Invigilator:")
        ws.cell(row=row_pointer, column=2, value=inv_map.get(room_name, ""))
        row_pointer += 2

        grid = data["grid"]

        # Column Headers
        for col in range(len(grid[0])):
            ws.cell(row=row_pointer, column=col + 2, value=f"Col {col+1}")

        row_pointer += 1

        # Rows
        for i, row in enumerate(grid):
            ws.cell(row=row_pointer, column=1, value=f"Row {i+1}")

            for j, seat in enumerate(row):
                ws.cell(row=row_pointer, column=j + 2, value=seat)

            row_pointer += 1

        row_pointer += 2

    return wb